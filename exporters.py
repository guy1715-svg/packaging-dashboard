"""
포장 사양 대시보드 - 내보내기(Export)
=======================================

구매팀 전달용 견적 요청서를 Excel / PDF 바이트로 생성합니다.
- Excel: openpyxl (구매팀이 '구매 확정 단가'를 직접 입력해 회신)
- PDF  : reportlab (한글 CID 폰트 사용)
"""

import io
from datetime import date

# 견적서 표 헤더 베트남어 병기 (성우비나용) ------------------------------------
VI_HEADERS = {
    "품명": "Tên SP", "박스종류": "Loại thùng", "박스명": "Tên thùng",
    "규격(Size)": "Kích thước", "포장재": "Vật liệu", "적재 방식": "Cách xếp",
    "박스당 총 제품": "SL/thùng", "제한 요인": "Giới hạn",
    "박스 총중량(kg)": "KL/thùng(kg)", "비고": "Ghi chú",
    "구매 확정 단가": "Đơn giá chốt",
}


def _col_label(col, bilingual):
    """열 이름 (병기 옵션이면 '한국어\\nTiếng Việt')."""
    if bilingual and col in VI_HEADERS:
        return f"{col}\n{VI_HEADERS[col]}"
    return col


# ---------------------------------------------------------------------------
# Excel 내보내기
# ---------------------------------------------------------------------------
def to_excel_bytes(header_info, rows, logo_bytes=None, chart_png=None,
                   bilingual=False):
    """
    header_info : dict (법인, 포장방식, 제품사이즈, 작성일 등)
    rows        : build_quote_rows() 결과 리스트
    logo_bytes  : 회사 로고 PNG bytes (있으면 상단 삽입)
    chart_png   : 3D 적재 배치도 PNG bytes (있으면 표 아래 삽입)
    bilingual   : True → 표 헤더 한국어+베트남어 병기
    반환        : xlsx 파일 bytes
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()
    ws = wb.active
    ws.title = "견적요청서"

    title_font = Font(size=14, bold=True)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2F5597")
    confirm_fill = PatternFill("solid", fgColor="FFF2CC")  # 구매 확정 단가 강조
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- 로고 (선택) ---
    if logo_bytes:
        try:
            img = XLImage(io.BytesIO(logo_bytes))
            img.height = min(img.height, 70)
            img.width = min(img.width, 200)
            ws.add_image(img, "F1")
        except Exception:
            pass

    # --- 제목 ---
    ws["A1"] = "포장자재 견적 요청서 (Packaging Quotation Request)"
    ws["A1"].font = title_font

    # --- 메타 정보 ---
    r = 3
    for k, v in header_info.items():
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        r += 1

    # --- 표 헤더 ---
    r += 1
    header_row = r
    if not rows:
        ws.cell(row=r, column=1, value="적재 가능한 박스가 없습니다. 제품 사이즈를 확인하세요.")
        return _wb_to_bytes(wb)

    columns = list(rows[0].keys())
    for c, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=c, value=_col_label(col, bilingual))
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = center
        cell.border = border

    # --- 데이터 ---
    confirm_col_idx = columns.index("구매 확정 단가") + 1 if "구매 확정 단가" in columns else None
    for i, row in enumerate(rows, start=header_row + 1):
        for c, col in enumerate(columns, start=1):
            val = row[col]
            cell = ws.cell(row=i, column=c, value=val)
            cell.alignment = center
            cell.border = border
            if confirm_col_idx and c == confirm_col_idx:
                cell.fill = confirm_fill  # 구매팀 입력란 강조

    # --- 열 너비 자동 ---
    for c, col in enumerate(columns, start=1):
        width = max(len(str(col)), 14) + 2
        ws.column_dimensions[ws.cell(row=header_row, column=c).column_letter].width = width

    # --- 안내 문구 ---
    note_row = header_row + len(rows) + 2
    ws.cell(row=note_row, column=1,
            value="※ 노란색 '구매 확정 단가' 열에 최종 단가를 입력 후 개발팀으로 회신 바랍니다.").font = Font(
        italic=True, color="C00000")

    # --- 3D 적재 배치도 이미지 (선택) ---
    if chart_png:
        try:
            cimg = XLImage(io.BytesIO(chart_png))
            cimg.width = min(cimg.width, 420)
            cimg.height = min(cimg.height, 320)
            ws.add_image(cimg, f"A{note_row + 2}")
        except Exception:
            pass

    return _wb_to_bytes(wb)


def _wb_to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF 내보내기 (한글 지원)
# ---------------------------------------------------------------------------
def to_pdf_bytes(header_info, rows, logo_bytes=None, chart_png=None,
                 bilingual=False):
    """구매팀 전달용 견적 요청서 PDF bytes 생성 (한글 CID 폰트).
    logo_bytes/chart_png 있으면 로고·3D 배치도 삽입, bilingual 이면 헤더 베트남어 병기."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, Image as RLImage)
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    # 한글 지원 내장 CID 폰트 등록
    font_name = "HYSMyeongJo-Medium"
    try:
        pdfmetrics.registerFont(UnicodeCIDFont(font_name))
    except Exception:
        font_name = "Helvetica"  # 폰트 미가용 시 폴백

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=14 * mm, bottomMargin=12 * mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"],
                                 fontName=font_name, fontSize=16)
    meta_style = ParagraphStyle("meta", parent=styles["Normal"],
                                fontName=font_name, fontSize=9, leading=13)
    note_style = ParagraphStyle("note", parent=styles["Normal"],
                                fontName=font_name, fontSize=8,
                                textColor=colors.HexColor("#C00000"))
    cell_style = ParagraphStyle("cell", parent=styles["Normal"],
                                fontName=font_name, fontSize=7, leading=9,
                                alignment=1)

    elems = []
    if logo_bytes:
        try:
            elems.append(RLImage(io.BytesIO(logo_bytes), width=44 * mm, height=16 * mm,
                                 kind="proportional"))
            elems.append(Spacer(1, 4))
        except Exception:
            pass
    elems += [Paragraph("포장자재 견적 요청서 (Packaging Quotation Request)", title_style),
              Spacer(1, 6)]

    meta_txt = "&nbsp;&nbsp;|&nbsp;&nbsp;".join(f"<b>{k}</b>: {v}" for k, v in header_info.items())
    elems.append(Paragraph(meta_txt, meta_style))
    elems.append(Spacer(1, 10))

    if not rows:
        elems.append(Paragraph("적재 가능한 박스가 없습니다. 제품 사이즈를 확인하세요.", meta_style))
        doc.build(elems)
        buf.seek(0)
        return buf.getvalue()

    columns = list(rows[0].keys())
    head_cell = ParagraphStyle("head", parent=cell_style, textColor=colors.white)
    data = [[Paragraph(_col_label(c, bilingual).replace("\n", "<br/>"), head_cell)
             for c in columns]]
    for row in rows:
        data.append([Paragraph("" if row[c] is None else str(row[c]), cell_style)
                     for c in columns])

    table = Table(data, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5597")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BFBFBF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F5FB")]),
    ]
    # '구매 확정 단가' 열 강조
    if "구매 확정 단가" in columns:
        ci = columns.index("구매 확정 단가")
        style.append(("BACKGROUND", (ci, 1), (ci, -1), colors.HexColor("#FFF2CC")))
    table.setStyle(TableStyle(style))
    elems.append(table)
    elems.append(Spacer(1, 10))
    elems.append(Paragraph(
        "※ 노란색 '구매 확정 단가' 열에 최종 단가를 기입 후 개발팀으로 회신 바랍니다.", note_style))

    if chart_png:
        try:
            elems.append(Spacer(1, 10))
            elems.append(Paragraph("■ 적재 배치도 (참고)", meta_style))
            elems.append(Spacer(1, 4))
            elems.append(RLImage(io.BytesIO(chart_png), width=110 * mm, height=82 * mm,
                                 kind="proportional"))
        except Exception:
            pass

    doc.build(elems)
    buf.seek(0)
    return buf.getvalue()


def default_header_info(customer, outer_group, product,
                        part_name="", unit_weight_g=0.0, inner_mode="", bag_name="",
                        sel_box="", sel_size="", sel_qty=0, tray_info=""):
    """견적서 상단 메타 정보 표준 생성 (고객사·선택 박스 포함)."""
    inner_label = inner_mode + (f" ({bag_name})" if (bag_name and "지퍼백" in inner_mode) else "")
    info = {
        "고객사": customer if customer else "-",
        "품명": part_name if part_name else "-",
        "포장재": inner_label if inner_mode else "-",
        "박스 종류": outer_group,
        "선택 박스": f"{sel_box} ({sel_size})" if sel_box else "-",
        "박스당 수량": f"{sel_qty:,} 개" if sel_qty else "-",
        "제품 사이즈(L×W×H mm)": f"{product[0]}×{product[1]}×{product[2]}",
        "제품 1개 무게(g)": f"{unit_weight_g:g}" if unit_weight_g else "-",
        "작성일": str(date.today()),
        "작성 부서": "개발팀 (R&D / Packaging)",
    }
    if tray_info:
        info["트레이 제작정보"] = tray_info
    return info
