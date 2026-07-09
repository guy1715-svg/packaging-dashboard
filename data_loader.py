"""
양산 엑셀(Master Data) 파싱 · 정제 스크립트
=============================================

양산처 엑셀(NEWBIZ 시트)에서 포장 핵심 데이터만 추출·정제하여
대시보드용 두 가지 DB로 변환합니다.

추출 컬럼 (NEWBIZ 시트, 헤더 3행 / 데이터 4행~)
  D(4)  Item              부품명
  F(6)  Code              부품 코드   ← 정답지 Key
  O(15) Packing Q'ty/Box  박스당 총수량
  Q(17) Carton Box        박스 코드명
  R(18) Size Box          박스 규격 (예: 400(L)*345(W)*410(H))

산출물
  ① PRODUCTION_BOXES : 고유 박스 규격 리스트 (대시보드 CARTONS 교체/확장용)
  ② PART_LOOKUP      : {부품코드: {item, carton, qty, L, W, H}} 검색용 dict

실행:  python data_loader.py <엑셀경로>   → master_data.py 생성
"""

import re
import sys
import openpyxl

SHEET = "NEWBIZ"
HEADER_ROW = 3
DATA_START = 4
COL = {"item": 4, "code": 6, "qty": 15, "carton": 17, "size": 18}


def clean_int(v):
    """'5250', '30ea', '160 Ea', '1,000' → 정수. 없으면 None."""
    if v is None:
        return None
    m = re.search(r"[\d,]+", str(v))
    if not m:
        return None
    try:
        return int(m.group(0).replace(",", ""))
    except ValueError:
        return None


def parse_size(v):
    """'400(L)*345(W)*410(H)' → (400, 345, 410). 숫자 3개 미만이면 None."""
    if v is None:
        return None
    nums = re.findall(r"\d+", str(v))
    if len(nums) < 3:
        return None
    return int(nums[0]), int(nums[1]), int(nums[2])


def clean_code(v):
    return str(v).strip() if v not in (None, "") else None


def clean_carton(v):
    return str(v).strip() if v not in (None, "") else None


def load_records(xlsx_path):
    """엑셀에서 정제된 레코드 리스트 반환."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET]
    records = []
    for r in range(DATA_START, ws.max_row + 1):
        code = clean_code(ws.cell(r, COL["code"]).value)
        if not code:
            continue
        carton = clean_carton(ws.cell(r, COL["carton"]).value)
        size = parse_size(ws.cell(r, COL["size"]).value)
        qty = clean_int(ws.cell(r, COL["qty"]).value)
        records.append({
            "code": code,
            "item": (str(ws.cell(r, COL["item"]).value).strip()
                     if ws.cell(r, COL["item"]).value else ""),
            "carton": carton,
            "qty": qty,
            "L": size[0] if size else None,
            "W": size[1] if size else None,
            "H": size[2] if size else None,
        })
    return records


def norm_carton(name):
    """박스명 정규화: 공백·대시 제거, 대문자, 끝의 BOX 제거 → 중복 병합용 키."""
    k = re.sub(r"[\s\-]+", "", str(name)).upper()
    k = re.sub(r"BOX$", "", k)
    return k


def build_box_db(records):
    """
    고유 박스 규격 DB. 정규화한 carton별로 대표 규격/이름을 뽑아 중복 병합.
    반환: [{박스명,size,inner_l,inner_w,inner_h,pack_type,재질,비고,box_cost,max_weight_kg,건수}, ...]
    """
    from collections import Counter, defaultdict
    sizes = defaultdict(Counter)     # 정규화키 → (L,W,H) Counter
    names = defaultdict(Counter)     # 정규화키 → 원본이름 Counter
    for rec in records:
        if rec["carton"] and rec["L"]:
            key = norm_carton(rec["carton"])
            sizes[key][(rec["L"], rec["W"], rec["H"])] += 1
            names[key][rec["carton"].strip()] += 1

    boxes = []
    for key, counter in sizes.items():
        (l, w, h), _ = counter.most_common(1)[0]
        disp = names[key].most_common(1)[0][0]          # 가장 많이 쓴 원본 이름
        boxes.append({
            "박스명": disp, "size": f"{l}*{w}*{h}",
            "inner_l": l, "inner_w": w, "inner_h": h,
            "pack_type": "box", "재질": "", "비고": "양산 실측",
            "box_cost": 0, "max_weight_kg": 20, "건수": sum(counter.values()),
        })
    boxes.sort(key=lambda b: -b["건수"])
    return boxes


def build_part_lookup(records):
    """부품코드 → {item, carton, qty, L, W, H} 검색용 dict (마지막 값 우선)."""
    lookup = {}
    for rec in records:
        lookup[rec["code"]] = {
            "item": rec["item"], "carton": rec["carton"], "qty": rec["qty"],
            "L": rec["L"], "W": rec["W"], "H": rec["H"],
        }
    return lookup


def write_master(xlsx_path, out_path="master_data.py"):
    """정제 결과를 정적 파이썬 파일로 저장 (앱에서 xlsx 없이 import 가능)."""
    records = load_records(xlsx_path)
    boxes = build_box_db(records)
    lookup = build_part_lookup(records)

    lines = ['"""양산 엑셀에서 자동 생성된 기준 데이터 (data_loader.py 산출물)."""', ""]
    lines.append("# ① 고유 박스 규격 (대시보드 CARTONS 교체/확장용 · data.py의 _box() dict와 동일 스키마)")
    lines.append("PRODUCTION_BOXES = [")
    for b in boxes:
        lines.append(
            f'    {{"박스명": {b["박스명"]!r}, "size": {b["size"]!r}, '
            f'"inner_l": {b["inner_l"]}, "inner_w": {b["inner_w"]}, "inner_h": {b["inner_h"]}, '
            f'"pack_type": "box", "재질": "", "비고": {b["비고"]!r}, '
            f'"box_cost": 0, "max_weight_kg": {b["max_weight_kg"]}}},  # {b["건수"]}건'
        )
    lines.append("]")
    lines.append("")
    lines.append("# ② 부품 코드 → 박스/수량 정답지")
    lines.append("PART_LOOKUP = {")
    for code, v in lookup.items():
        lines.append(f'    {code!r}: {v!r},')
    lines.append("}")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    return records, boxes, lookup


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "master.xlsx"
    recs, boxes, lookup = write_master(path)
    print(f"레코드 {len(recs)}건 · 고유 박스 {len(boxes)}종 · 부품 {len(lookup)}개")
    print("master_data.py 생성 완료")
